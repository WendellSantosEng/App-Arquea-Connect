"""
Nome:
    Código de Supressão - Safra 21-23
    Código desenvolvido em Pt-br para fins de trabalho

Sobre o autor:
    Nome: Wendell Resende dos Santos
    Cargo: Estagiário na empresa ARQUEA Engenharia e Geotecnologia
    GitHub: https://github.com/WendellSantosEng
"""

# Importação dos módulos/bibliotecas nescessárias
# - pathlib para encontrar a Planilha nescessária
# - openpyxl para desenvolver utilizando planilhas .xlsx
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

def codigo_supressao(df):

    # Criar um novo Workbook
    # Criar um novo Workbook
    workbook = Workbook()
    
    # Adicionar um novo Worksheet
    worksheet = workbook.active
    worksheet.title = "Codigos Car"

    # Preencher o Worksheet com os dados do DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False)):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx + 1, column=c_idx, value=value)

    # Nome para a planilha
    new_sheet_name = 'Codigo supressao'

    # Verifica se a aba já existe
    if new_sheet_name not in workbook.sheetnames:
        # Adiciona a nova aba
        new_worksheet = workbook.create_sheet(new_sheet_name)
    else:
        # Seleciona a aba existente
        new_worksheet = workbook[new_sheet_name]

    # Seleciona a aba existente (Aba 1)
    origin_sheet_name = 'Codigos Car'
    origin_worksheet = workbook[origin_sheet_name]

    # Criando index com base em colunas
    first_column = origin_worksheet.min_column
    list_columns = []
    for i in range(first_column, first_column + 13):
        list_columns.append(i)

    list_names = []
    dict_car = { }
    list_cars = []

    # Concertando as celulas da worksheet original
    for row in origin_worksheet.iter_rows():
        for cell in row:
            # Evitando quaisquer erros que podem ter vindo na planilha, esses erros são os mais comuns
            value = str(cell.value)
            value = value.replace(",", "")
            value = value.replace("&amp;lt;", "<").replace("&lt;", "<").replace("&amp;", "<").replace("< ", "<")
            value = value.replace("ã", "a").replace("ç", "c")
            cell.number_format = "General"
            cell.value = '' if cell.value is None else value

            # Criando o Dict com key= CAR, value= Repetição desse mesmo, na planilha
            if cell.column == list_columns[2]:
                if cell.value is not None and cell.value != "" and cell.value != "None" and cell.row != 1:
                    list_cars.append(cell.value)
                    if cell.value not in dict_car:
                        dict_car[cell.value] = 1
                    else:
                        dict_car[cell.value] += 1

    # Cria uma lista
    codigos = []

    # Gerando códigos exclusivos para cada entrada na lista de CAR
    # Essa lista tera o valor da repetição do CAR na lista e a enumeração que ele aparece
    # EXEMPLO: CAR aparece 15 vezes, e esta é a sua 4° ocorrencia, logo: 0015-04
    for nome in list_cars:
        count = str(dict_car[nome]).zfill(4)  # Garante pelo menos 4 dígitos no count, preenchendo com zeros à esquerda
        index = 1
        while (f"{count}-{index:02}") in codigos:  # Verifica se o código já existe
            index += 1
        codigos.append(f"{count}-{index:02}")

    # Cria strings que serão preenchidas ao decorrer da Planilha
    coluna12 = ""
    year_cult = ""

    # PRINCIPAL FUNÇÃO DO PROGRAMA
    # Passa por cada celula e maneja seu conteudo
    for row, codigo in zip(origin_worksheet.iter_rows(min_row=2), codigos):

        # Cria uma string vazia, a qual sera apendada na lista final sendo assim o código exclusivo para cada supressão, gerado
        line = ""

        for cell in row:

            # Verifica se o valor não é None(Nulo)
            if cell.value is not None:
                content = cell.value

                # Repetição de codigo CAR na tabela seguido de seu índice de ocorrência
                if cell.column == list_columns[2]:
                    # Adicionando o código correspondente ao CAR
                    line += codigo
                    line += "-"

                # Adiciona a letra C, e logo após os ultimos 5 digitos do Codigo CAR
                if cell.column == list_columns[2]:
                    line += "C"
                    car5dig = str(cell.value)[-5:]
                    line += car5dig
                    line += "-"

                # Faz uma verificação precisa se o conteúdo da célula é Pós 2017 (ou suas variações) sendo assim é adicionado
                # D17, o que significa Depois de 2017.
                # Caso o valor Pré 2017 (ou suas variações) o valor adicionado será A17, o que quer dizer, Antes de 2017
                if cell.column == list_columns[3]:
                    y = cell.value
                    if y == "PÓS 2017" or y == "POS2017" or y == "PÓS2017" or y == "POS 2017":
                        line += "D17"
                    elif y == "PRÉ 2017" or y == "PRE2017" or y == "PRÉ2017" or y == "PRE 2017":
                        line += "A17"
                    line += "-"

                # O código faz com que se o valor for <2017 (Antes de 2017) seja adicionado a string, o valor 00
                # Caso o valor seja diferente disso coloca os últimos digitos do ano em questão
                if cell.column == list_columns[12]:
                    coluna12 = ""
                    yr = cell.value
                    yr = yr.replace(" ", "")
                    if yr == "<2017":
                        coluna12 += "00"
                    else:
                        ano = str(yr[-2:])
                        coluna12 += ano

                # Captura a primeira letra da célula
                if cell.column == list_columns[8]:
                    words = str(cell.value).split()
                    first_char = [word[0] for word in words]
                    line += "".join(first_char)
                
                # Verificar se o conteúdo da célula é "CULTIVO" ou nao
                # Caso seja coloca os ultimos dois digitos do ano, o qual é o nome da coluna
                # Caso não seja, adiciona o valor "-00-"
                # Isso se repete pras próximas 3 verificações de if
                if cell.column == list_columns[5]: # 2021
                    year_cult = ""
                    new_value = str(cell.value)
                    new_value = new_value.replace(" ","")
                    if new_value == "CULTIVO":
                        year_cult += "-21-"
                    else:
                        year_cult += "-00-"

                if cell.column == list_columns[6]: # 2021
                    new_value = str(cell.value)
                    new_value = new_value.replace(" ","")
                    if new_value == "CULTIVO":
                        year_cult += "22-"
                    else:
                        year_cult += "00-"

                if cell.column == list_columns[7]: # 2021
                    new_value = str(cell.value)
                    new_value = new_value.replace(" ","")
                    if new_value == "CULTIVO":
                        year_cult += "23"
                    else:
                        year_cult += "00"

        # Nesse momento, é onde é adicionado o conteúdo de algumas strings que captaram valores durante a execução
        # Além de adicionar a string formada (line) em uma List
        if line:
            line = line[:19] + coluna12 + "-" + line[19:]
            line += year_cult
            list_names.append(line)

    # Aqui o código recorre a um procedimento que faz com que na nova planilha gerada, a primeira coluna dela seja cada item dessa lista
    # Adicione os valores nas células da nova coluna
    for index, value in enumerate(list_names, start=2):
        new_worksheet.cell(row=index, column=1, value=value)

    # Corrigindo formatações da coluna gerada, e do seu conteúdo passado
    row:tuple[Cell]
    for row in new_worksheet.iter_rows():
        for cell in row:
            value = str(cell.value)
            value = value.replace("&amp;lt;", "<")
            value = value.replace("&lt;", "<")
            value = value.replace("&amp;", "<")
            cell.value = value
            if cell.row == 1 and cell.column == 1:
                cell.value = "COD SUP"

    # Transita as informações dessa coluna para a primeira coluna da Planilha original
    for row in range(2, new_worksheet.max_row + 1):
        value = new_worksheet[f'A{row}'].value
        origin_worksheet[f'A{row}'] = value

    # Apaga a nova Planilha que foi gerada
    worksheet_name_to_remove = 'Codigo supressao'
    if worksheet_name_to_remove in workbook.sheetnames:
        worksheet_to_remove = workbook[worksheet_name_to_remove]
        workbook.remove(worksheet_to_remove)

    for row in range(2, new_worksheet.max_row + 1):
        value = new_worksheet[f'A{row}'].value
        origin_worksheet[f'A{row}'] = value

    return workbook