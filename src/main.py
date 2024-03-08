"""

Autor: Wendell Resende dos Santos
Data de Criação: 15/01/2024 ~ 08/03/2024

Descrição:
Esta aplicação foi desenvolvida para realização de serviços de geoprocessamento da empresa Arquea Engenharia e Geotecnologia

Uso:
Uso exclusivo da equipe Arquea

"""


import sys
import io
import os
from typing import cast
import threading
import time
import datetime
import csv
import qdarkstyle
from PySide6.QtCore import QCoreApplication, QPropertyAnimation, QEasingCurve, QObject, Signal, QUrl
from PySide6.QtGui import QIcon, QDesktopServices
from PySide6.QtWidgets import (QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox)
import pandas as pd
import geopandas as gpd
from shapely.geometry import Polygon
from shapely.wkt import loads
import shapely
from codsup import codigo_supressao
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from mainwindow import Ui_MainWindow


class ShapefileLoader(QObject):
    progress_updated = Signal(int)
    shapefile_loaded = Signal(gpd.GeoDataFrame)

    def __init__(self):
        super().__init__()
        self._stop_event = threading.Event()

    def load_shapefile(self, sicar_path, column_names):
        try:
            # Lendo o arquivo Shapefile como GeoDataFrame
            sicarmt = gpd.read_file(sicar_path)
            
            # Definindo os nomes das colunas
            sicarmt.columns = column_names

            # Verificar os nomes das colunas antes e depois da atribuição
            print("Nomes das colunas antes da atribuição:", sicarmt.columns)
            
            self.shapefile_loaded.emit(sicarmt)

            # Verificar algumas informações sobre o GeoDataFrame após a atribuição
            print("Informações sobre o GeoDataFrame após a atribuição:")
            print(sicarmt.head())
            print(sicarmt.info())

            # Definindo o valor máximo da barra de progresso como o número total de registros no GeoDataFrame
            total_records = len(sicarmt)
            self.progress_updated.emit(total_records)

            # Atualizando a barra de progresso conforme os registros são processados
            for i in range(total_records):
                # Simulando o processamento de cada registro (você pode substituir isso pelo processamento real)
                time.sleep(0.01)
                self.progress_updated.emit(i + 1)

                # Verificando se a thread deve ser encerrada
                if self._stop_event.is_set():
                    break

                # Processa os eventos do aplicativo regularmente para garantir que a interface do usuário não congele
                QCoreApplication.processEvents()

            return sicarmt

        except Exception as e:
            print(f"Erro ao carregar o Shapefile: {e}")

    def stop_loading(self):
        # Define o evento de parada para interromper o processamento
        self._stop_event.set()

class ClassShape(QObject):

    def __init__(self, ui_main_window):
        super().__init__()
        self.ui_main_window = ui_main_window

    def start_load_shapefile(self, loader, sicar_path, column_names):
        loader.load_shapefile(sicar_path, column_names)

    def update_progress(self, value):
        self.ui_main_window.progressBar_sicar.setValue(value)
        if value == self.ui_main_window.progressBar_sicar.maximum():
            self.ui_main_window.label_aguarde.hide()

    def handle_shapefile_loaded(self, sicarmt):
        print("carregou")
        # Verificando se o GeoDataFrame foi carregado corretamente
        if not sicarmt.empty:
            print("O GeoDataFrame não está vazio.")
        else:
            print("O GeoDataFrame está vazio.")

        # Exibindo algumas informações sobre o GeoDataFrame
        print("Informações sobre o GeoDataFrame:")
        print(sicarmt.head())
        print(sicarmt.info())

        # Este método é chamado quando o shapefile foi carregado
        # Você pode fazer qualquer coisa que precise com o sicarmt aqui
        # Por exemplo, você pode atribuí-lo a uma variável de instância para acessá-lo posteriormente
        self.ui_main_window.sicarmt = sicarmt
        print("ta carregado")
        self.ui_main_window.btn_selectCAR.setEnabled(True)

class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.setWindowTitle("ArqueaConnect")
        appIcon = QIcon(u"")
        self.setWindowIcon(appIcon)
        self.stackedWidget.setCurrentWidget(self.page_home)

        # Criar um novo ShapefileLoader
        self.loader = ShapefileLoader()
        self.loader.progress_updated.connect(self.update_progress)
        self.loader.shapefile_loaded.connect(self.handle_shapefile_loaded)

        # Botao do menu lateral
        self.pushButton.clicked.connect(self.leftMenu)

        self.toolBox_usinas.setCurrentIndex(-1) 

        # Paginas do sistema
        self.pushButton_2.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page_home))
        self.pushButton_9.clicked.connect(lambda: (
            self.stackedWidget_3.setCurrentWidget(self.page),
            self.stackedWidget.setCurrentWidget(self.page_shape)
        ))
        self.pushButton_5.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.page_sobre))

        # Botões para entrar em funcionalidades de cada usina
        self.pushButton_11.clicked.connect(lambda: self.stackedWidget_3.setCurrentWidget(self.page_5))
        self.pushButton_8.clicked.connect(lambda: self.stackedWidget_3.setCurrentWidget(self.page_2))
        self.pushButton_10.clicked.connect(lambda: self.stackedWidget_3.setCurrentWidget(self.page_4))
        self.pushButton_3.clicked.connect(lambda: self.stackedWidget_3.setCurrentWidget(self.page_3))

        # Botões de voltar
        self.btn_back_1.clicked.connect(lambda: (self.stackedWidget_3.setCurrentWidget(self.page)))
        self.btn_back_2.clicked.connect(lambda: (self.stackedWidget_3.setCurrentWidget(self.page)))
        self.btn_back_3.clicked.connect(lambda: (self.stackedWidget_3.setCurrentWidget(self.page)))
        self.btn_back_LRV_1.clicked.connect(lambda: (self.stackedWidget_3.setCurrentWidget(self.page)))
        self.btn_back_LRV_3.clicked.connect(lambda: (self.stackedWidget_3.setCurrentWidget(self.page)))
        self.btn_back_SUP.clicked.connect(lambda: (self.stackedWidget_3.setCurrentWidget(self.page)))

        # Variáveis para armazenar os dados do shapefile
        self.gdf = None
        self.gdf_atributos = None

        # Botao para abrir arquivos Excel
        self.btn_import.clicked.connect(lambda: self.abrir_arquivos("",None, None, self.tableWidget))

        # Botao para abrir arquivos Shapefile
        self.btn_importshp.clicked.connect(self.abrir_shp)

        # Botão para fazer modificações
        self.btn_gerar.clicked.connect(self.gerar_codigo_sup)

        # Botão para salvar modificações
        self.btn_saved.clicked.connect(self.salvar_modificacoes)

        # Botão para salvar Shapefile
        self.btn_savedshp.clicked.connect(lambda: self.exportar_shp_da_tabela(self.tableWidget, "SUP", None))

        # Botao para abrir Planilha de Dados        
        self.btn_planilha_dados.clicked.connect(lambda: self.abrir_arquivos("DATA", self.tableWidget_2, self.tableWidget_LRV_dados))
        self.btn_planilha_dados_LRV.clicked.connect(lambda: self.abrir_arquivos("DATA", self.tableWidget_2, self.tableWidget_LRV_dados))
        
        self.df_SRS = None
        self.df_LRV = None

        # Botão para entrar no filtro
        self.filtered_rows_SRS = []
        self.btn_filtrar.clicked.connect(lambda: self.show_filter_menu(self.tableWidget_2, self.frame_filter, self.select_column))
        self.frame_filter.hide()

        self.filtered_rows_LRV = []
        self.btn_filtrar_LRV.clicked.connect(lambda: self.show_filter_menu(self.tableWidget_LRV_dados, self.frame_filter_2, self.select_column_LRV))
        self.frame_filter_2.hide()

        # Botão pra aparecer a tabela de acordo com o filtro
        self.column_index_SRS = -1
        self.column_index_LRV = -1
        self.select_column.currentIndexChanged.connect(lambda: setattr(self, 'column_index_SRS', self.column_selected(self.select_column)))
        self.select_column_LRV.currentIndexChanged.connect(lambda: setattr(self, 'column_index_LRV', self.column_selected(self.select_column_LRV)))
        
        self.filtered_rows = {}  # Dicionário para armazenar os índices das linhas filtradas para cada tabela

        # Conecte a função intermediária ao sinal clicked do botão
        self.btn_okfilter.clicked.connect(lambda: self.filtred(self.search_filter, self.tableWidget_2, self.column_index_SRS))
        self.btn_okfilter_LRV.clicked.connect(lambda: self.filtred(self.search_filter_LRV, self.tableWidget_LRV_dados, self.column_index_LRV))

        # Importando SICAR-MT
        self.btn_sicar.clicked.connect(self.abrir_sicarmt)

        # Desabilitar o botão btn_selectCAR inicialmente
        self.btn_selectCAR_SRS.setEnabled(False)
        self.btn_selectCAR_LRV.setEnabled(False)

        # Selecionando as linhas de dentro do SicarMT de acordo com o que é visivel na tabela de Dados principal
        self.btn_selectCAR_SRS.clicked.connect(lambda: self.seletionCARs(self.df_SRS))
        self.btn_selectCAR_LRV.clicked.connect(lambda: self.seletionCARs(self.df_LRV))
        self.label_aguarde.hide()
        self.progressBar_sicar.hide()
        self.progressBar_sicar.setRange(0, 100)
        self.column_names = []

        # Exportando como Shp a tabela do SicarMT
        self.gdfSICAR = None
        self.gdf_atributosSICAR = None
        self.shp_car_name = None
        self.shp_car_year = None
        self.name_general_safra = None
        self.btn_exportCAR.clicked.connect(lambda: self.exportar_shp_da_tabela(self.tableWidget_3, "CAR", self.shp_car_name))

        self.shp_safra_name_SRS = None
        self.shp_safra_name_LRV = None
        self.exportSHPSafra.clicked.connect(lambda: self.exportar_shp_da_tabela(self.tableWidget_4, "SAFRA", self.shp_safra_name_SRS))
        self.exportSHPSafra_LRV.clicked.connect(lambda: self.exportar_shp_da_tabela(self.tableWidget_LRVsafra, "SAFRA", self.shp_safra_name_LRV))

        self.importDataSafra.clicked.connect(lambda: self.import_data_safra_func(self.df_SRS, self.tableWidget_4))
        self.importDataSafra_LRV.clicked.connect(lambda: self.import_data_safra_func(self.df_LRV, self.tableWidget_LRVsafra))
        
        self.car_values_safra_SRS = []
        self.car_values_safra_LRV = []
        self.btn_CSV.clicked.connect(lambda: self.openCsvFile(self.car_values_safra_SRS, self.tableWidget_4))
        self.btn_CSV2.clicked.connect(lambda: self.openCsvFile(self.car_values_safra_LRV, self.tableWidget_LRVsafra))

        self.btn_arquea.clicked.connect(self.open_website_arquea)
        self.btn_github.clicked.connect(self.open_website_github)

    def open_website_arquea(self):
        url = "https://arqueageotec.com.br/"
        QDesktopServices.openUrl(QUrl(url))

    def open_website_github(self):
        url = "https://github.com/WendellSantosEng/App-Arquea-Connect"
        QDesktopServices.openUrl(QUrl(url))

    def gerar_codigo_sup(self):
        # Obter os dados da tabela widget
        dados = []
        cabecalhos = []

        for column in range(self.tableWidget.columnCount()):
            cabecalhos.append(self.tableWidget.horizontalHeaderItem(column).text())

        for row in range(self.tableWidget.rowCount()):  # Começar da primeira linha
            linha = []
            for column in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, column)
                if item is not None:
                    linha.append(item.text())
                else:
                    linha.append("")
            dados.append(linha)

        # Criar um DataFrame do pandas com os dados da tabela e os cabeçalhos
        df = pd.DataFrame(dados, columns=cabecalhos)

        # Inicializar list_geometry com uma lista vazia
        list_geometry = []

        # Verificar se a coluna "geometry" está presente no DataFrame
        if 'geometry' in df.columns:
            list_geometry = df['geometry'].tolist()
            df.drop(columns=['geometry'], inplace=True)
        else:
            print("A coluna 'geometry' não está presente no DataFrame.")

        # Chamar a função para modificar os dados
        workbook_modificado = codigo_supressao(df)
        ws = workbook_modificado.active

        # Verificar se a coluna 'geometry' já existe
        if 'geometry' not in df.columns:
            last_column_index = len(df.columns) + 1
            ws.insert_cols(idx=last_column_index)

            # Adicionar os dados da lista à coluna "geometry"
            for idx, item in enumerate(list_geometry, start=2):
                ws.cell(row=idx, column=last_column_index, value=item)

            # Renomear a coluna para 'geometry'
            ws.cell(row=1, column=last_column_index, value='geometry')
        else:
            print("A coluna 'geometry' já existe.")

        # Atualizar a QTableWidget com os dados do workbook
        self.mostrar_dados(workbook_modificado, self.tableWidget)

    # Menu da esquerda
    def leftMenu(self):
        width = self.left_container.width()
        print(width)

        if width == 9:
            newWidth = 150
        else:
            newWidth = 9

        self.animation = QPropertyAnimation(self.left_container, b"maximumWidth")
        self.animation.setDuration(500)
        self.animation.setStartValue(width)
        self.animation.setStartValue(width)
        self.animation.setEndValue(newWidth)
        self.animation.setEasingCurve(QEasingCurve.InOutQuart)
        self.animation.start()
    
    def mostrar_dados_CBX(self, worksheet, tablewidget):
        if not isinstance(worksheet, Worksheet):
            print("Erro: O objeto não é uma instância válida de Worksheet")
            return

        dados = []
        for linha in worksheet.iter_rows(values_only=True, min_row=2):
            if not all(cell is None for cell in linha):
                dados.append(linha)

        if dados:
            # Exiba os dados na tabela widget
            tablewidget.setRowCount(len(dados))
            tablewidget.setColumnCount(len(HEADERS_DATA))
            for i, linha in enumerate(dados):
                for j, valor in enumerate(linha):
                    item = QTableWidgetItem(str(valor))
                    tablewidget.setItem(i, j, item)

            # Adicione os cabeçalhos das colunas
            tablewidget.setHorizontalHeaderLabels(HEADERS_DATA)

    def abrir_arquivos(self, value, tablewidgetSRS: QTableWidget = None, tablewidgetLRV: QTableWidget = None, tablewidget: QTableWidget = None):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo Excel", "", "Arquivos Excel (*.xlsx *.xls)", options=options)
        if fileName:
            workbook = load_workbook(fileName)
            if value == "DATA":
                if "SRS - ENTRADA DE DADOS" in workbook.sheetnames:
                    worksheet_srs = workbook["SRS - ENTRADA DE DADOS"]
                    self.mostrar_dados_CBX(worksheet_srs, tablewidgetSRS)
                if "LRV - ENTRADA DE DADOS" in workbook.sheetnames:
                    worksheet_lrv = workbook["LRV - ENTRADA DE DADOS"]
                    self.mostrar_dados_CBX(worksheet_lrv, tablewidgetLRV)
            else:
                if tablewidget is not None:
                    # Mostrar dados na tabela especificada
                    self.mostrar_dados(workbook, tablewidget)

    def parse_coordinates(self, coord_str):
        coords = []
        # Dividir a string em substrings individuais com base na vírgula
        for coord_pair in coord_str.split(','):
            # Verificar se a substring não contém a palavra 'POLYGON'
            if 'POLYGON' not in coord_pair:
                # Remover os parênteses extras e converter a substring em um par de coordenadas
                coord_pair = coord_pair.strip().strip(')').strip('(')
                coords.append(tuple(map(float, coord_pair.split())))
        return coords
    
    def transformar_coluna_data_str(self, workbook):
        worksheet = workbook.active
        data_column_index = None
        for col_idx, col in enumerate(worksheet.iter_cols(values_only=True)):
            if col[0] == "DATA":
                data_column_index = col_idx
                break
        if data_column_index is not None:
            for i, cell in enumerate(worksheet.iter_rows(values_only=True, min_row=2)):
                if isinstance(cell[data_column_index], datetime.datetime):
                    # Converte a data para o formato desejado dia;mes;ano
                    data_str = cell[data_column_index].strftime("%d/%m/%Y")
                    # Atualiza o valor na célula
                    worksheet.cell(row=i + 2, column=data_column_index + 1, value=data_str)
        else:
            print("A coluna 'DATA' não foi encontrada.")

    def abrir_shp(self):
        shp_path, _ = QFileDialog.getOpenFileName(self, "Selecionar Shapefile", "", "Arquivos Shapefile (*.shp)")

        if shp_path:
            # Lendo o arquivo Shapefile como GeoDataFrame
            gdf = gpd.read_file(shp_path)

            # Convertendo as geometrias para strings
            gdf['geometry'] = gdf['geometry'].apply(lambda geom: str(geom))

            # Inicializar um novo workbook do Excel
            wb = Workbook()
            ws = wb.active

            # Adicionar os dados do GeoDataFrame ao workbook do Excel
            for row in dataframe_to_rows(gdf, index=False, header=True):
                ws.append(row)

            # Exibir os dados na tabelaWidget
            self.mostrar_dados(wb, self.tableWidget)

    # Método para mostrar dados na tabela widget e atualizar a planilha
    def mostrar_dados(self, workbook, tablewidget):
        # Atualizar a QTableWidget com os dados do workbook
        worksheet = workbook.active
        dados = []
        for linha in worksheet.iter_rows(values_only=True, min_row=2):
            if not all(cell is None for cell in linha):
                dados.append(linha)

        df = pd.DataFrame(dados, columns=[coluna[0].value for coluna in worksheet.iter_cols()])
        
        tablewidget.setRowCount(df.shape[0])
        tablewidget.setColumnCount(df.shape[1])
        tablewidget.setHorizontalHeaderLabels(df.columns)
        
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                if df.columns[j] == "DATA":  # Verifica se a coluna é "DATA"
                    # Se for, converte o valor para string
                    tablewidget.setItem(i, j, QTableWidgetItem(str(df.iloc[i, j])))
                else:
                    # Caso contrário, insere o valor normalmente
                    tablewidget.setItem(i, j, QTableWidgetItem(str(df.iloc[i, j])))

    # Método para salvar as modificações na planilha original
    def salvar_modificacoes(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getSaveFileName(self, "Salvar Tabela como Excel", "", "Arquivo Excel (*.xlsx)")

        if fileName:
            # Criar um DataFrame a partir dos dados da tableWidget
            dados = []
            cabecalhos = []
            for column in range(self.tableWidget.columnCount()):
                cabecalhos.append(self.tableWidget.horizontalHeaderItem(column).text())

            for row in range(self.tableWidget.rowCount()):
                linha = []
                for column in range(self.tableWidget.columnCount()):
                    item = self.tableWidget.item(row, column)
                    if item is not None:
                        linha.append(item.text())
                    else:
                        linha.append("")
                dados.append(linha)

            df = pd.DataFrame(dados, columns=cabecalhos)

            # Salvar o DataFrame como um arquivo Excel
            df.to_excel(fileName, index=False)

    def show_filter_menu(self, tablewidget, framefilter, select_column):
        # Verifica se há uma tabela no tableWidget_2
        if tablewidget.rowCount() > 0 and tablewidget.columnCount() > 0:
            # Limpa as opções anteriores do combobox
            select_column.clear()
            # Preenche as opções do combobox com os nomes das colunas da tabela
            for col in range(tablewidget.columnCount()):
                column_name = tablewidget.horizontalHeaderItem(col).text()
                if column_name == "Qual o código de identificação do Grupo?":
                    column_name = "ID"
                select_column.addItem(column_name)
            # Exibe o frame do filtro
            framefilter.show()
        else:
            # Se não houver tabela, oculta o frame do filtro
            framefilter.hide()

    def column_selected(self, select_column):
        # Retorna o índice da coluna selecionada no combobox
        return select_column.currentIndex()

    def filtred(self, search_filter, tablewidget, column_index):
        # Obtém o texto de filtro do lineedit
        filter_text = search_filter.text().lower()
        print(filter_text, " em ", column_index)

        if tablewidget is self.tableWidget_2:
            # DataFrame para armazenar os dados filtrados
            self.df_SRS = pd.DataFrame(columns=[str(tablewidget.horizontalHeaderItem(col).text()) for col in range(tablewidget.columnCount())])
        if tablewidget is self.tableWidget_LRV_dados:
            # DataFrame para armazenar os dados filtrados
            self.df_LRV = pd.DataFrame(columns=[str(tablewidget.horizontalHeaderItem(col).text()) for col in range(tablewidget.columnCount())])
            
        # Verifica se um índice de coluna válido foi selecionado
        # Verifica se um índice de coluna válido foi selecionado
        if column_index >= 0:
            # Itera sobre as linhas da tabela
            for row in range(tablewidget.rowCount()):
                # Variável para verificar se a linha está vazia
                is_row_empty = True
                
                # Variável para verificar se a linha atende aos critérios de filtro
                passes_filter = False
                
                # Itera sobre as colunas da linha atual
                for col in range(tablewidget.columnCount()):
                    # Verifica se há um item não vazio na célula
                    item = tablewidget.item(row, col)
                    if item and item.text():
                        is_row_empty = False
                        # Verifica se o texto da célula atende ao critério de filtro
                        if filter_text in item.text().lower():
                            passes_filter = True
                            break  # Se encontrar um item que atende ao filtro, sai do loop
                
                # Exibe a linha se atender aos critérios de filtro e não estiver vazia
                if not is_row_empty and passes_filter:
                    tablewidget.setRowHidden(row, False)
                    # Adiciona os dados da linha ao DataFrame correspondente
                    row_data = []
                    for col in range(tablewidget.columnCount()):
                        if not tablewidget.isColumnHidden(col):
                            row_data.append(tablewidget.item(row, col).text())
                    if tablewidget is self.tableWidget_2:
                        self.df_SRS.loc[len(self.df_SRS)] = row_data
                    elif tablewidget is self.tableWidget_LRV_dados:
                        self.df_LRV.loc[len(self.df_LRV)] = row_data
                else:
                    # Oculta a linha se não atender aos critérios de filtro ou estiver vazia
                    tablewidget.setRowHidden(row, True)
        else:
            # Se nenhum índice de coluna válido foi selecionado, exibe todas as linhas da tabela
            for row in range(tablewidget.rowCount()):
                tablewidget.setRowHidden(row, False)
        
        # Restaura o estado filtrado anterior, se existir
        self.restore_filter(tablewidget)

        NOME = "Qual o nome do Grupo a ser informado?"
        ANO = "Qual o ano da safra analisada?"

        if self.df_SRS is not None:
            self.shp_safra_name_SRS = self.df_SRS[NOME].iloc[0]
            self.shp_safra_name_SRS = self.shp_safra_name_SRS.replace(" ", "_")
            year = self.df_SRS[ANO].iloc[0]
            self.shp_safra_name_SRS = self.shp_safra_name_SRS + "_" + str(year)

        if self.df_LRV is not None:
            self.shp_safra_name_LRV = self.df_LRV[NOME].iloc[0]
            self.shp_safra_name_LRV = self.shp_safra_name_LRV.replace(" ", "_")
            year = self.df_LRV[ANO].iloc[0]
            self.shp_safra_name_LRV = self.shp_safra_name_LRV + "_" + str(year)

        if tablewidget is self.tableWidget_2:
            self.name_general_safra = self.shp_safra_name_SRS
            print("DENTRO SRS: ", self.name_general_safra)
        elif tablewidget is self.tableWidget_LRV_dados:
            self.name_general_safra = self.shp_safra_name_LRV
            print("DENTRO LRV: ", self.name_general_safra)
        
        # Imprime o DataFrame
        print("SRS: ",self.df_SRS)
        print("LRV: ",self.df_LRV)

    def restore_filter(self, tablewidget):
        # Restaura o filtro anterior, se existir, para a tabela atual
        if tablewidget in self.filtered_rows:
            filtered = self.filtered_rows[tablewidget]
            for row in range(tablewidget.rowCount()):
                if row not in filtered:
                    tablewidget.setRowHidden(row, True)

    def abrir_sicarmt(self):
        sicar_path, _ = QFileDialog.getOpenFileName(None, "Selecionar Shapefile", "", "Arquivos Shapefile (*.shp)")

        if sicar_path:
            print("CARREGOU")
            self.progressBar_sicar.setValue(0)
            self.label_aguarde.show()
            self.progressBar_sicar.show()

            # Obter os nomes das colunas da tabela
            self.column_names = [self.tableWidget_3.horizontalHeaderItem(column).text() for column in range(self.tableWidget_3.columnCount())]

            # Conectar os sinais e slots do loader
            self.loader.progress_updated.connect(self.update_progress)
            self.loader.shapefile_loaded.connect(self.handle_shapefile_loaded)

            # Iniciar a carga do shapefile na thread
            self.loader.load_shapefile(sicar_path, self.column_names)
        
    def update_progress(self, value):
        self.progressBar_sicar.setValue(value)
        if value == self.progressBar_sicar.maximum():
            self.label_aguarde.hide()

    def handle_shapefile_loaded(self, sicarmt):
        # Verificando se o GeoDataFrame foi carregado corretamente
        if not sicarmt.empty:
            print("O GeoDataFrame não está vazio.")
        else:
            print("O GeoDataFrame está vazio.")

        # Exibindo algumas informações sobre o GeoDataFrame
        print("Informações sobre o GeoDataFrame:")
        print(sicarmt.head())
        print(sicarmt.info())
        
        # Este método é chamado quando o shapefile foi carregado
        # Você pode fazer qualquer coisa que precise com o sicarmt aqui
        # Por exemplo, você pode atribuí-lo a uma variável de instância para acessá-lo posteriormente
        self.sicarmt = sicarmt
        self.btn_selectCAR_LRV.setEnabled(True)
        self.btn_selectCAR_SRS.setEnabled(True)

    def seletionCARs(self, df):
        CAR = "Informe os códigos de CARs do Grupo"
        NOME = "Qual o nome do Grupo a ser informado?"
        ANO = "Qual o ano da safra analisada?"

        if not isinstance(df, pd.DataFrame):
            df = pd.DataFrame(df)
        
        # Verifica se a coluna CAR está presente no DataFrame
        if CAR in df.columns:

            self.shp_car_name = df[NOME].iloc[0]
            self.shp_car_year = df[ANO].iloc[0]
            self.shp_car_name = self.shp_car_name.replace(" ","_")
            self.shp_car_name = "CAR_" + self.shp_car_name + "_" + self.shp_car_year
            print()
            print("CAR NAME: ",self.shp_car_name)
            print()

            car_values_df = df[CAR].unique()  # Obtém os valores únicos da coluna CAR

            # Filtra as linhas do GeoDataFrame sicarmt que possuem valores presentes na coluna CAR do DataFrame
            selected_rows = self.sicarmt[self.sicarmt['cod_imovel'].isin(car_values_df)]

            # Limpa a tabelaWidget_3
            self.tableWidget_3.clear()
            
            # Define o número de linhas e colunas na tabelaWidget_3
            self.tableWidget_3.setRowCount(selected_rows.shape[0])
            self.tableWidget_3.setColumnCount(selected_rows.shape[1])

            # Define os cabeçalhos da tabelaWidget_3 usando a lista 'headers'
            self.tableWidget_3.setHorizontalHeaderLabels(HEADER_CAR)
            
            # Preenche a tabelaWidget_3 com os dados do DataFrame selected_rows
            for i, row in enumerate(selected_rows.iterrows()):
                for j, value in enumerate(row[1]):
                    item = QTableWidgetItem(str(value))
                    self.tableWidget_3.setItem(i, j, item)
        else:
            print("A coluna 'CAR' não está presente no DataFrame.")

    def extrair_dados_da_tabela(self, tablewidget):
        data = []
        header_items = [tablewidget.horizontalHeaderItem(column).text() for column in range(tablewidget.columnCount())]
        
        for row in range(tablewidget.rowCount()):
            row_data = []
            for column in range(tablewidget.columnCount()):
                cell_item = tablewidget.item(row, column)
                if cell_item is not None:
                    row_data.append(cell_item.text())
                else:
                    row_data.append("")  
            data.append(row_data)
            
        return data, header_items
    
    def makedir_shp(self, new_folder_path):
        os.mkdir(new_folder_path)

    def exportar_shp_da_tabela(self, tablewidget, value, nome_shp_safra: str | None):
        if value == "CAR":
            print("NOME: ", nome_shp_safra)
            data = []
            for row in range(tablewidget.rowCount()):
                row_data = []
                for column in range(tablewidget.columnCount()):
                    cell_item = tablewidget.item(row, column)
                    if cell_item is not None:
                        row_data.append(cell_item.text())
                    else:
                        row_data.append("")  
                data.append(row_data)

            # Pergunta ao usuário se deseja criar as pastas necessárias
            resposta = QMessageBox.question(self, "Criar Pastas", "Deseja criar as pastas necessárias?",
                                            QMessageBox.Yes | QMessageBox.No)
            
            # Se o usuário clicou em 'Sim', cria as pastas
            if resposta == QMessageBox.Yes:
                # Seleciona o diretório onde os arquivos serão salvos
                shp_dir = QFileDialog.getExistingDirectory(self, "Selecione o diretório para salvar Shapefile")
                
                if shp_dir:
                    # Cria a estrutura de diretórios necessária
                    shp_folder = os.path.join(shp_dir, nome_shp_safra[4:])
                    car_dir = os.path.join(shp_folder, f"CAR_{nome_shp_safra[-4:]}")
                    safra_dir = os.path.join(shp_folder, f"SAFRA_{nome_shp_safra[-4:]}")

                    if not os.path.exists(shp_folder):
                        os.mkdir(shp_folder)
                    if not os.path.exists(car_dir):
                        os.mkdir(car_dir)
                    if not os.path.exists(safra_dir):
                        os.mkdir(safra_dir)
                    
                    car_name = nome_shp_safra
                    # Caminho completo para salvar o arquivo na pasta CAR
                    shp_path = os.path.join(car_dir, f"{car_name}.shp")
                    
                    df = pd.DataFrame(data, columns=HEADER_CAR)
                    df['geometry'] = None
                    gdf = gpd.GeoDataFrame(df, geometry=[shapely.geometry.Polygon() for _ in range(len(df))])
                    gdf.crs = 'EPSG:4674'
                    gdf.to_file(shp_path, driver='ESRI Shapefile')

                    if 'polygon' in df.columns:

                        print("TEM COLUNA POLYGON")
                        try:
                            df['geometry'] = df['polygon'].apply(lambda coords: shapely.geometry.Polygon(self.parse_coordinates(coords)) if isinstance(coords, str) else None)

                            # Remover linhas com geometria inválida
                            df = df.dropna(subset=['geometry'])

                            # Remover a coluna "polygon" agora que a coluna de geometria está presente
                            df.drop(columns=['polygon'], inplace=True)

                            # Criar um GeoDataFrame a partir do DataFrame 'df'
                            gdf = gpd.GeoDataFrame(df)
                            gdf.crs = 'EPSG:4674'
                        except NameError:
                            print("Erro ao criar geometria. Verifique se o pacote shapely está corretamente importado.")
                            return
                    else:
                        return
                    
                    if 'polygon' in df.columns:
                        print("TEM A COLUNA AINDA")
                        
                    gdf.to_file(shp_path, driver='ESRI Shapefile')

            # Se o usuário clicou em 'Não', pede apenas para selecionar a pasta de destino
            elif resposta == QMessageBox.No:
                print("NOME NAO: ", nome_shp_safra)
                shp_path, _ = QFileDialog.getSaveFileName(self, "Salvar Shapefile", nome_shp_safra, "Arquivos Shapefile (*.shp)")
                if shp_path:
                    df = pd.DataFrame(data, columns=HEADER_CAR)
                    df['geometry'] = None
                    gdf = gpd.GeoDataFrame(df, geometry=[shapely.geometry.Polygon() for _ in range(len(df))])
                    gdf.to_file(shp_path, driver='ESRI Shapefile')

                if 'polygon' in df.columns:
                    try:
                        df['geometry'] = df['polygon'].apply(lambda coords: shapely.geometry.Polygon(self.parse_coordinates(coords)) if isinstance(coords, str) else None)

                        # Remover linhas com geometria inválida
                        df = df.dropna(subset=['geometry'])

                        # Remover a coluna "polygon" agora que a coluna de geometria está presente
                        df.drop(columns=['polygon'], inplace=True)

                        # Criar um GeoDataFrame a partir do DataFrame 'df'
                        gdf = gpd.GeoDataFrame(df)
                        gdf.crs = 'EPSG:4674'
                    except NameError:
                        print("Erro ao criar geometria. Verifique se o pacote shapely está corretamente importado.")
                        return
                else:
                    return
                    
                gdf.to_file(shp_path, driver='ESRI Shapefile')
                
        elif value == "SAFRA":
            def replace_empty_string(value):
                if value == '':
                    return None
                return value
            dtype_list = [str, str,str,str,int,str,str,str,str,str,str,float,float,str,str,]
            data, header_items = self.extrair_dados_da_tabela(tablewidget)
            data = [row for row in data if all(item is not None for item in row)]  # Remover linhas com células vazias
            # Pergunta ao usuário se deseja criar as pastas necessárias
            resposta = QMessageBox.question(self, "Criar Pastas", "Deseja criar as pastas necessárias?",
                                            QMessageBox.Yes | QMessageBox.No)
            
            # Se o usuário clicou em 'Sim', cria as pastas
            if resposta == QMessageBox.Yes:
                # Seleciona o diretório onde os arquivos serão salvos
                shp_dir = QFileDialog.getExistingDirectory(self, "Selecione o diretório para salvar Shapefile")
                
                if shp_dir:
                    # Cria a estrutura de diretórios necessária
                    shp_folder = os.path.join(shp_dir, nome_shp_safra)
                    car_dir = os.path.join(shp_folder, f"CAR_{nome_shp_safra[-4:]}")
                    safra_dir = os.path.join(shp_folder, f"SAFRA_{nome_shp_safra[-4:]}")
                    
                    if not os.path.exists(shp_folder):
                        os.mkdir(shp_folder)
                    if not os.path.exists(car_dir):
                        os.mkdir(car_dir)
                    if not os.path.exists(safra_dir):
                        os.mkdir(safra_dir)
                    
                    # Caminho completo para salvar o arquivo na pasta CAR
                    shp_path = os.path.join(safra_dir, f"{nome_shp_safra}.shp")
                    
                    df = pd.DataFrame(data, columns=header_items)
                    df = df.applymap(replace_empty_string)
                    
                    for i, dtype in enumerate(dtype_list):
                        if i < len(df.columns):  # Verificar se o índice está dentro do intervalo das colunas do DataFrame
                            df[df.columns[i]] = df[df.columns[i]].astype(dtype)
                    df['geometry'] = None
                    gdf = gpd.GeoDataFrame(df, geometry=[shapely.geometry.Polygon() for _ in range(len(df))])
                    gdf.crs = 'EPSG:4674'
                    gdf.to_file(shp_path, driver='ESRI Shapefile')

            # Se o usuário clicou em 'Não', pede apenas para selecionar a pasta de destino
            elif resposta == QMessageBox.No:
                shp_path, _ = QFileDialog.getSaveFileName(self, "Salvar Shapefile", nome_shp_safra, "Arquivos Shapefile (*.shp)")
                if shp_path:
                    df = pd.DataFrame(data, columns=header_items)
                    df = df.applymap(replace_empty_string)
                    for i, dtype in enumerate(dtype_list):
                        if i < len(df.columns):  # Verificar se o índice está dentro do intervalo das colunas do DataFrame
                            df[df.columns[i]] = df[df.columns[i]].astype(dtype)
                    df['geometry'] = None
                    gdf = gpd.GeoDataFrame(df, geometry=[shapely.geometry.Polygon() for _ in range(len(df))])
                    gdf.crs = 'EPSG:4674'
                    gdf.to_file(shp_path, driver='ESRI Shapefile')

        elif value == "SUP":
            data, header_items = self.extrair_dados_da_tabela(tablewidget)
            shp_path, _ = QFileDialog.getSaveFileName(None, "Salvar Shapefile", "", "Arquivos Shapefile (*.shp)")
            if shp_path:
                df = pd.DataFrame(data, columns=header_items)

                if 'polygon' in df.columns:
                    try:
                        df['geometry'] = None
                        gdf = gpd.GeoDataFrame(df, geometry=[shapely.geometry.Polygon() for _ in range(len(df))])
                        gdf.crs = 'EPSG:4674'
                        gdf.to_file(shp_path, driver='ESRI Shapefile')

                        # df['geometry'] = df['polygon'].apply(lambda coords: shapely.geometry.Polygon(self.parse_coordinates(coords)) if isinstance(coords, str) else None)

                        # # Remover linhas com geometria inválida
                        # df = df.dropna(subset=['geometry'])

                        # # Remover a coluna "polygon" agora que a coluna de geometria está presente
                        # df.drop(columns=['polygon'], inplace=True)

                        # # Criar um GeoDataFrame a partir do DataFrame 'df'
                        # gdf = gpd.GeoDataFrame(df)

                    except NameError:
                        print("Erro ao criar geometria. Verifique se o pacote shapely está corretamente importado.")
                        return
                else:
                    return
                    
                gdf.to_file(shp_path, driver='ESRI Shapefile')

        else:
            print("Valor não reconhecido para exportação.")

    def import_data_safra_func(self, df, tablewidget):
        
        columns_mapping = {
            "O Grupo pertence a qual usina?": 1,
            "Qual o código de identificação do Grupo?": 0,
            "Qual o nome do Grupo a ser informado?": 2,
            "Qual o CPF/CNPJ do Grupo": 3,
            "Qual o ano da safra analisada?": 4,
            "Informe os códigos de CARs do Grupo": 6,
            "DATA": 9,
        }

        headers_from_dict = list(columns_mapping.keys())

        # Limpando tableWidget antes de copiar
        tablewidget.clear()
        tablewidget.setRowCount(0)

        headers = [
            "COD_PRODUT", "PLANT_USIN", "NOM_PRODUT", "CPF_CNPJ", "ANO_SAFRA", "TIPO", "COD_IMOVEL", "COD_MAPA", "CIDADE",
            "DATA_REGIS", "SITUACAO", "AREA_TOTAL", "AREA_MAPA", "SUPRESSAO", "ELEGIBILID",
        ]

        tablewidget.clear()
        tablewidget.setColumnCount(len(headers))
        tablewidget.setHorizontalHeaderLabels(headers)
        
        if tablewidget is self.tableWidget_4:
            self.car_values_safra_SRS.clear
        elif tablewidget is self.tableWidget_LRVsafra:
            self.car_values_safra_LRV.clear

        # Transferir os dados do DataFrame para tableWidget_4
        for row_position, (_, row_data) in enumerate(df.iterrows()):
            for source_col_name, target_col_index in columns_mapping.items():
                try:
                    # Obter o valor da coluna no DataFrame
                    value = row_data[source_col_name]
                    # Se a linha correspondente não existir em tableWidget_4, crie-a
                    if row_position >= tablewidget.rowCount():
                        tablewidget.insertRow(row_position)
                    # Definir o valor na célula correspondente em tableWidget_4
                    tablewidget.setItem(row_position, target_col_index, QTableWidgetItem(str(value)))

                    # Verificar se o nome da coluna é "DATA" e formatar seu valor
                    if source_col_name == "DATA":
                        data_y_m_d = datetime.datetime.strptime(value, "%d/%m/%Y").strftime("%Y-%m-%d")
                        tablewidget.setItem(row_position, target_col_index, QTableWidgetItem(data_y_m_d))
                    
                    # Adicionar o valor da coluna "COD_IMOVEL" à lista
                    if source_col_name == "Informe os códigos de CARs do Grupo":
                        if tablewidget is self.tableWidget_4:
                            self.car_values_safra_SRS.append(value)
                        elif tablewidget is self.tableWidget_LRVsafra:
                            self.car_values_safra_LRV.append(value)

                except KeyError as e:
                    print(f"Erro: Coluna não encontrada no DataFrame - {e}")
                except ValueError as e:
                    print(f"Erro: {e}")

        # Exemplo de obtenção do nome da usina
        column_index = headers_from_dict.index("O Grupo pertence a qual usina?")
        if tablewidget.rowCount() > 0:
            name_usina = tablewidget.item(0, column_index).text()

    def openCsvFile(self, list_, tablewidget):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo CSV", "", "Arquivos CSV (*.csv)", options=options)
        if fileName:
            try:
                # Aumentar o limite de tamanho do campo permitido pelo sistema operacional (apenas para sistemas Unix-like)
                if sys.platform.startswith('linux') or sys.platform.startswith('darwin'):
                    csv.field_size_limit(sys.maxsize)  # Definir o tamanho máximo do campo

                # Leia apenas as colunas de índice 1 e 2
                columns_to_read = ['registro_no_car', 'data_de_cadastro']  # Índices das colunas que deseja ler

                data_for_car_values = {}
                chunksize = 10000  # Tamanho do chunk (ajuste conforme necessário)

                # Leia o arquivo CSV em pedaços menores (chunks)
                for chunk in pd.read_csv(fileName, usecols=columns_to_read, chunksize=chunksize, encoding='latin1', sep=';'):
                    # Itere sobre os pedaços do DataFrame
                    for _, row in chunk.iterrows():
                        car = row.iloc[0]  # A primeira coluna
                        data = row.iloc[1]  # A segunda coluna
                        data = data[:10] # Tira a hora e fica só data
                        if car in list_:
                            if car not in data_for_car_values:
                                data_for_car_values[car] = []
                            data_for_car_values[car].append(data)

                # Supondo que sua QTableWidget seja chamada de tableWidget
                for car, data_list in data_for_car_values.items():
                    for row_index in range(tablewidget.rowCount()):
                        item = tablewidget.item(row_index, 6)  # Supondo que a coluna 'car' seja a coluna 0
                        if item and item.text() == car:
                            # Se encontrarmos a correspondência do 'car' na linha da tabela,
                            # pegamos a data correspondente da lista de datas e a inserimos na tabela
                            first_date_string = data_list[0]
                            first_date = datetime.datetime.strptime(first_date_string, '%d/%m/%Y').strftime('%Y-%m-%d')
                            # Convertendo de volta para objeto datetime
                            first_date_datetime = datetime.datetime.strptime(first_date, '%Y-%m-%d')
                            tablewidget.setItem(row_index, 9, QTableWidgetItem(first_date))  # Supondo que a coluna 'data' seja a coluna 1

                # Exibir dados para o usuário
                message = ""
                for car, data in data_for_car_values.items():
                    message += f"Dados para {car}: {data}\n"
                print("Dados de CAR: ", message)
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao abrir o arquivo: {e}")
                print(f"{e}")

    def closeEvent(self, event):
        event.accept()
        sys.exit()

if __name__ == '__main__':

    HEADER_CAR = [
        "cod_imovel", "status_imo", "dat_criaca", "area", "condicao", "uf", "municipio", "cod_munic", "m_fiscal",
        "tipo_imove", "polygon",
    ]
    HEADERS_DATA = [
        "Carimbo de data/hora", "Nome do analista CBX", "O Grupo pertence a qual usina?", "Qual o código de identificação do Grupo?", 
        "Qual o nome do Grupo a ser informado?", "Qual o CPF/CNPJ do Grupo", "Será enviado shape do Grupo pela CBX?", 
        "Qual o ano da safra analisada?", "Qual o tipo de processo realizado?", "Observações referente ao processo realizado" , 
        "Qual a área declarada do Grupo?", "Informe a quantidade de CARs pertencente ao Grupo", 
        "Informe os códigos de CARs do Grupo", "Endereço de e-mail" , "__PowerAppsId__",
    ]

    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec())

